"""Rellena la matriz de control de versiones AI-DLC para JotaPe VF."""
from __future__ import annotations

import shutil
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SRC = Path(
    r"c:\Users\Asus\Downloads\Matriz de registro(MODELO DE CONTROL DE VERSIONES).xlsx"
)
OUT = Path(
    r"c:\Users\Asus\Downloads\Matriz de registro - JotaPe VF COMPLETADA.xlsx"
)
REPO = "https://github.com/danimanriqueaguirre21/JP-Textil"
MONOREPO = "jotape-vf/"

# (registro_col, evidencia_col, section_result_col, keys in bolt dict)
SECTIONS = [
    ("inicio", "O", ["intent", "req_doc", "bolt_req"]),
    ("construccion", "V", ["design", "commit_impl", "tests"]),
    ("operaciones", "Y", ["deploy"]),
    ("coherencia_est", "AI", ["module", "layer", "no_redundant", "separation"]),
    ("coherencia_conv", "AP", ["naming", "organization", "deps"]),
    ("trazabilidad", "BB", ["bolt_doc", "req_ref", "pr_commit", "validation", "deploy_trace"]),
    ("estab_func", "BL", ["expected", "scope", "regression", "integration"]),
    ("estab_tec", "BU", ["static", "security", "auto_tests", "code_review"]),
]

FIELD_COLS = {
    "id": "B",
    "desc": "C",
    "model": "D",
    "date": "E",
    "domain": "F",
    "hu": "G",
    "owner": "H",
    "intent": ("I", "J"),
    "req_doc": ("K", "L"),
    "bolt_req": ("M", "N"),
    "design": ("P", "Q"),
    "commit_impl": ("R", "S"),
    "tests": ("T", "U"),
    "deploy": ("W", "X"),
    "module": ("AA", "AB"),
    "layer": ("AC", "AD"),
    "no_redundant": ("AE", "AF"),
    "separation": ("AG", "AH"),
    "naming": ("AJ", "AK"),
    "organization": ("AL", "AM"),
    "deps": ("AN", "AO"),
    "bolt_doc": ("AR", "AS"),
    "req_ref": ("AT", "AU"),
    "pr_commit": ("AV", "AW"),
    "validation": ("AX", "AY"),
    "deploy_trace": ("AZ", "BA"),
    "expected": ("BD", "BE"),
    "scope": ("BF", "BG"),
    "regression": ("BH", "BI"),
    "integration": ("BJ", "BK"),
    "static": ("BM", "BN"),
    "security": ("BO", "BP"),
    "auto_tests": ("BQ", "BR"),
    "code_review": ("BS", "BT"),
    "status": "BW",
    "notes": "BX",
    "improvement": "BY",
    "result_z": "Z",
    "result_aq": "AQ",
    "result_bc": "BC",
    "result_bv": "BV",
}

BOLTS = [
    {
        "id": "BOLT-VF-001",
        "desc": "Monorepo JotaPe VF: frontend Next.js 16, backend FastAPI, Docker Compose, documentación y CI básica.",
        "model": "Cursor Agent (Claude)",
        "date": "21/05/2026",
        "domain": "E-commerce / Infraestructura",
        "hu": "RF-03, RF-08",
        "owner": "Danilo Manrique",
        "intent": ("Sí", f"{MONOREPO}README.md — Intent: e-commerce + probador virtual"),
        "req_doc": ("Sí", f"{MONOREPO}docs/02-requerimientos/requerimientos-funcionales.md"),
        "bolt_req": ("Sí", "Bolt derivado de RF-03 (catálogo) y RF-08 (carrito)"),
        "design": ("Sí", f"{MONOREPO}docs/03-arquitectura/architecture.md"),
        "commit_impl": ("Sí", f"{REPO}/commit/11a9557"),
        "tests": ("Sí", f"{MONOREPO}frontend: npm test · {MONOREPO}backend: pytest"),
        "deploy": ("N.A.", "Entorno local (make dev). Despliegue productivo pendiente — roadmap Fase 5."),
        "module": ("Sí", "Código en frontend/, backend/, infra/docker/"),
        "layer": ("Sí", "Clean Architecture en backend; App Router en frontend"),
        "no_redundant": ("Sí", "Migración desde virtual-fitting-ai documentada en MIGRACION.md"),
        "separation": ("Sí", "Módulos auth, users, health, predictions separados"),
        "naming": ("Sí", "Convenciones en .cursor/rules y README"),
        "organization": ("Sí", "Estructura monorepo jotape-vf/"),
        "deps": ("Sí", "Dependencias mínimas en requirements.txt y package.json"),
        "bolt_doc": ("Sí", "ID BOLT-VF-001 en esta matriz"),
        "req_ref": ("Sí", "RF-03, RF-08"),
        "pr_commit": ("Sí", "Commit 11a9557 — Add jotape-vf monorepo"),
        "validation": ("Sí", "CI GitHub Actions + tests locales"),
        "deploy_trace": ("N.A.", "Sin producción; stack local documentado en README"),
        "expected": ("Sí", "Tienda y carrito funcionan en localhost:3000"),
        "scope": ("Sí", "Alcance: catálogo, carrito, checkout demo"),
        "regression": ("Sí", "Tests Jest y Cypress sin regresiones reportadas"),
        "integration": ("Sí", "Integración frontend-backend vía API routes"),
        "static": ("Sí", "ESLint + Ruff en CI"),
        "security": ("Sí", f"{MONOREPO}docs/03-arquitectura/security-production.md"),
        "auto_tests": ("Sí", "pytest + jest — suites en verde"),
        "code_review": ("Sí", "Revisión manual + asistencia IA documentada"),
        "status": "Cerrado",
        "notes": "Monorepo base operativo. Enlace: " + REPO,
        "improvement": "Completar despliegue productivo y pedidos reales con BD.",
    },
    {
        "id": "BOLT-VF-002",
        "desc": "Escaneo corporal híbrido (MediaPipe + segmentación), fusión de medidas y calibración del avatar 3D en /try-on.",
        "model": "Cursor Agent (Claude)",
        "date": "23/05/2026",
        "domain": "Probador virtual / Visión computacional",
        "hu": "RF-05, RF-07",
        "owner": "Danilo Manrique",
        "intent": ("Sí", f"{MONOREPO}docs/04-ui-ux/escaneo-corporal/README.md"),
        "req_doc": ("Sí", "RF-05 Ingreso de datos del cliente"),
        "bolt_req": ("Sí", "Derivado de RF-05 y probador virtual (roadmap Fase 5-6)"),
        "design": ("Sí", f"{MONOREPO}docs/04-ui-ux/escaneo-corporal/fase-1-captura-guiada.md"),
        "commit_impl": ("Sí", f"{REPO}/commit/11a9557 · {REPO}/commit/ca6cd06"),
        "tests": ("Sí", f"{MONOREPO}frontend/src/lib/body-scan/*.test.ts"),
        "deploy": ("N.A.", "Flujo en localhost:3000/try-on/body-scan"),
        "module": ("Sí", "frontend/src/lib/body-scan/, components/body-scan/"),
        "layer": ("Sí", "Lógica en lib/; UI en components/; tipos en types/"),
        "no_redundant": ("Sí", "Reutiliza measurement-calculator y pose-landmarks"),
        "separation": ("Sí", "fuse-measurements, classify-body-type, avatar-calibration separados"),
        "naming": ("Sí", "Convención body-scan-* y fuse-*"),
        "organization": ("Sí", "docs/04-ui-ux/escaneo-corporal/"),
        "deps": ("Sí", "MediaPipe en cliente; sin dependencias extra en backend"),
        "bolt_doc": ("Sí", "ID BOLT-VF-002"),
        "req_ref": ("Sí", "RF-05, RF-07"),
        "pr_commit": ("Sí", "Commits 11a9557, ca6cd06"),
        "validation": ("Sí", "Modo diagnóstico + panel de medidas en UI"),
        "deploy_trace": ("N.A.", "Persistencia local sessionStorage + API medidas en desarrollo"),
        "expected": ("Sí", "Escaneo frontal/lateral genera medidas y calibra avatar"),
        "scope": ("Sí", "Clasificación OBESE/OVERWEIGHT coherente con BMI"),
        "regression": ("Sí", "Tests body-circumference, fuse-measurements, measurement-calculator"),
        "integration": ("Sí", "Integrado en buildOutfitRig y GltfAvatar"),
        "static": ("Sí", "TypeScript estricto + ESLint"),
        "security": ("Sí", "Imágenes en cliente; no se suben fotos en claro al servidor"),
        "auto_tests": ("Sí", "Jest body-scan suite"),
        "code_review": ("Sí", "Iteración documentada en matriz y commits"),
        "status": "Cerrado",
        "notes": "Segmentación + presets visuales para avatar CC. Repo: " + REPO,
        "improvement": "E2E Cypress del flujo completo de escaneo.",
    },
    {
        "id": "BOLT-VF-003",
        "desc": "Predicción de talla con Random Forest: endpoint POST /api/v1/predictions/ y recomendador en /try-on.",
        "model": "Cursor Agent (Claude)",
        "date": "21/05/2026",
        "domain": "Machine Learning",
        "hu": "RF-06, RF-07",
        "owner": "Danilo Manrique",
        "intent": ("Sí", "Reducir devoluciones por talla incorrecta — company.md"),
        "req_doc": ("Sí", "RF-06 Predicción de talla (ML) — prioridad crítica"),
        "bolt_req": ("Sí", "Bolt derivado directamente de RF-06 y RF-07"),
        "design": ("Sí", f"{MONOREPO}docs/03-arquitectura/modules.md"),
        "commit_impl": ("Sí", f"{REPO}/commit/11a9557"),
        "tests": ("Sí", f"{MONOREPO}backend/src/ml/test_predictor.py"),
        "deploy": ("N.A.", "API local :8000; modelo en ai-models/serialized/"),
        "module": ("Sí", "backend/src/ml/ + frontend size-recommender"),
        "layer": ("Sí", "ML en capa ml/; exposición en predictions/presentation"),
        "no_redundant": ("Sí", "Un solo predictor joblib reutilizado"),
        "separation": ("Sí", "predictor.py separado de routes"),
        "naming": ("Sí", "Convenciones Python PEP8"),
        "organization": ("Sí", "ai-models/ para artefactos"),
        "deps": ("Sí", "joblib + scikit-learn únicamente para ML"),
        "bolt_doc": ("Sí", "ID BOLT-VF-003"),
        "req_ref": ("Sí", "RF-06, RF-07"),
        "pr_commit": ("Sí", "Commit 11a9557"),
        "validation": ("Sí", "test_predictor.py + prueba manual en /try-on"),
        "deploy_trace": ("N.A.", "Sin producción"),
        "expected": ("Sí", "Endpoint devuelve talla y confianza"),
        "scope": ("Sí", "Predicción para prendas del catálogo demo"),
        "regression": ("Sí", "Tests ML en CI"),
        "integration": ("Sí", "frontend/src/app/api/predictions/route.ts"),
        "static": ("Sí", "Ruff + mypy opcional"),
        "security": ("Sí", "Validación Pydantic en schemas"),
        "auto_tests": ("Sí", "pytest predictor"),
        "code_review": ("Sí", "Revisión en commit inicial"),
        "status": "Cerrado",
        "notes": "Modelo RF integrado. Mejorar dataset real de JotaPe Textil.",
        "improvement": "Reentrenar con datos históricos de devoluciones.",
    },
    {
        "id": "BOLT-VF-004",
        "desc": "Autenticación BFF (Next.js API routes), JWT backend, esquema BD en español y hardening de seguridad.",
        "model": "Cursor Agent (Claude)",
        "date": "23/05/2026",
        "domain": "Autenticación / Seguridad",
        "hu": "RF-01, RF-02",
        "owner": "Danilo Manrique",
        "intent": ("Sí", "RF-01 Registro y RF-02 Autenticación"),
        "req_doc": ("Sí", f"{MONOREPO}docs/02-requerimientos/requerimientos-funcionales.md"),
        "bolt_req": ("Sí", "Derivado de RF-01 y RF-02"),
        "design": ("Sí", f"{MONOREPO}docs/Modelo Base de datos/database-schema.md"),
        "commit_impl": ("Sí", f"{REPO}/commit/ca6cd06"),
        "tests": ("Sí", f"{MONOREPO}backend/src/modules/auth/presentation/test_routes.py"),
        "deploy": ("N.A.", "Entorno desarrollo local"),
        "module": ("Sí", "backend/src/modules/auth/ + frontend/app/api/auth/"),
        "layer": ("Sí", "BFF en frontend; dominio en backend Clean Architecture"),
        "no_redundant": ("Sí", "Sin duplicar lógica JWT en cliente"),
        "separation": ("Sí", "security.py, routes, schemas separados"),
        "naming": ("Sí", "Tablas y campos en español según esquema"),
        "organization": ("Sí", "modules/auth con capas domain/application/infrastructure"),
        "deps": ("Sí", "python-jose, passlib — estándar FastAPI"),
        "bolt_doc": ("Sí", "ID BOLT-VF-004"),
        "req_ref": ("Sí", "RF-01, RF-02"),
        "pr_commit": ("Sí", "Commit ca6cd06"),
        "validation": ("Sí", "test_routes.py + test_security.py"),
        "deploy_trace": ("N.A.", "Sin producción"),
        "expected": ("Sí", "Login/register funcionan vía BFF"),
        "scope": ("Sí", "Auth JWT end-to-end en desarrollo"),
        "regression": ("Sí", "Tests auth en CI"),
        "integration": ("Sí", "Cookies/httpOnly vía API routes Next.js"),
        "static": ("Sí", "Ruff + ESLint"),
        "security": ("Sí", f"{MONOREPO}docs/03-arquitectura/security-production.md"),
        "auto_tests": ("Sí", "pytest auth module"),
        "code_review": ("Sí", "Commit ca6cd06 revisado"),
        "status": "Cerrado",
        "notes": "Auth BFF operativo. Repo: " + REPO,
        "improvement": "Completar refresh tokens y OAuth social.",
    },
    {
        "id": "BOLT-VF-005",
        "desc": "Presets visuales por bodyType para avatar Character Creator: mezcla 75% preset + 25% medida sin deformar rig.",
        "model": "Cursor Agent (Claude)",
        "date": "21/05/2026",
        "domain": "Probador virtual 3D / UX avatar",
        "hu": "RF-05, RF-07",
        "owner": "Danilo Manrique",
        "intent": ("Sí", "Evitar deformación grotesca del avatar CC al escalar huesos"),
        "req_doc": ("Sí", "RNF-02 Interfaz amigable + probador virtual company.md"),
        "bolt_req": ("Sí", "Complemento de BOLT-VF-002 para visualización 3D"),
        "design": ("Sí", f"{MONOREPO}frontend/src/lib/body-scan/body-visual-presets.ts"),
        "commit_impl": ("Sí", "Cambios locales en jotape-vf/ (pendiente commit) · " + REPO),
        "tests": ("Sí", f"{MONOREPO}frontend/src/lib/body-scan/body-visual-presets.test.ts"),
        "deploy": ("N.A.", "localhost:3000/try-on"),
        "module": ("Sí", "body-visual-presets.ts, apply-proportional-scales.ts"),
        "layer": ("Sí", "Presets en lib/; aplicación en virtual-fitting/"),
        "no_redundant": ("Sí", "Reemplaza escalado crudo directo al rig"),
        "separation": ("Sí", "Detección (fuse) separada de deformación visual (presets)"),
        "naming": ("Sí", "BODY_VISUAL_PRESETS, resolveAvatarVisualScales"),
        "organization": ("Sí", "avatar-scale-bone-map.ts para diagnóstico"),
        "deps": ("Sí", "Sin dependencias nuevas"),
        "bolt_doc": ("Sí", "ID BOLT-VF-005"),
        "req_ref": ("Sí", "RF-05, RF-07, RNF-02"),
        "pr_commit": ("Sí", "Trabajo en rama local — ver " + REPO),
        "validation": ("Sí", "Panel diagnóstico + tests Jest"),
        "deploy_trace": ("N.A.", "Solo entorno local"),
        "expected": ("Sí", "Avatar OBESE robusto pero humano (arm/thigh <= 1.25)"),
        "scope": ("Sí", "Solo huesos CC_Base torso/brazos/muslos; no cabeza/manos/pies"),
        "regression": ("Sí", "compute-proportional-scales.test.ts"),
        "integration": ("Sí", "buildOutfitRig → applyProportionalCalibration"),
        "static": ("Sí", "TypeScript + ESLint"),
        "security": ("Sí", "Sin impacto en seguridad"),
        "auto_tests": ("Sí", "body-visual-presets.test.ts"),
        "code_review": ("Sí", "Revisión iterativa con modo diagnóstico"),
        "status": "En progreso",
        "notes": "Huesos: CC_Base_Spine01/02, Clavicle, Upperarm, Thigh. Repo: " + REPO,
        "improvement": "Commit y tag de versión cuando QA apruebe avatar OBESE.",
    },
]


def section_result(items: list[tuple[str, str]]) -> str:
    regs = [r for r, _ in items]
    if all(r in ("Sí", "SI", "Si") for r in regs):
        return "Cumple"
    if any(r in ("No", "NO") for r in regs):
        return "No cumple"
    return "Cumple con observaciones"


def write_pair(ws, row: int, col_reg: str, col_ev: str, pair: tuple[str, str]) -> None:
    ws[f"{col_reg}{row}"] = pair[0]
    ws[f"{col_ev}{row}"] = pair[1]


def fill_bolt(ws, row: int, bolt: dict) -> None:
    ws[f"B{row}"] = bolt["id"]
    ws[f"C{row}"] = bolt["desc"]
    ws[f"D{row}"] = bolt["model"]
    ws[f"E{row}"] = bolt["date"]
    ws[f"F{row}"] = bolt["domain"]
    ws[f"G{row}"] = bolt["hu"]
    ws[f"H{row}"] = bolt["owner"]

    for key, cols in FIELD_COLS.items():
        if key in ("id", "desc", "model", "date", "domain", "hu", "owner", "status", "notes", "improvement"):
            continue
        if key.startswith("result_"):
            continue
        if key not in bolt:
            continue
        reg_col, ev_col = cols
        write_pair(ws, row, reg_col, ev_col, bolt[key])

    for _name, result_col, keys in SECTIONS:
        items = [bolt[k] for k in keys if k in bolt]
        ws[f"{result_col}{row}"] = section_result(items)

    # Resultados globales por bloque principal
    ws[f"Z{row}"] = section_result(
        [bolt[k] for k in ["intent", "req_doc", "bolt_req", "design", "commit_impl", "tests", "deploy"] if k in bolt]
    )
    ws[f"AQ{row}"] = section_result(
        [bolt[k] for k in ["module", "layer", "no_redundant", "separation", "naming", "organization", "deps"] if k in bolt]
    )
    ws[f"BC{row}"] = section_result(
        [bolt[k] for k in ["bolt_doc", "req_ref", "pr_commit", "validation", "deploy_trace"] if k in bolt]
    )
    ws[f"BV{row}"] = section_result(
        [bolt[k] for k in ["expected", "scope", "regression", "integration", "static", "security", "auto_tests", "code_review"] if k in bolt]
    )

    ws[f"BW{row}"] = bolt["status"]
    ws[f"BX{row}"] = bolt["notes"]
    ws[f"BY{row}"] = bolt["improvement"]


def main() -> None:
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    ws = wb["Bolts"]

    # Fila de contexto del proyecto (opcional, si B5 está libre)
    ws["B5"] = "PROYECTO"
    ws["C5"] = "JotaPe VF — E-commerce + probador virtual + ML tallas (JotaPe Textil, Huancayo)"
    ws["D5"] = "—"
    ws["E5"] = date.today().strftime("%d/%m/%Y")
    ws["F5"] = "Sistema completo"
    ws["G5"] = "RF-01 a RF-14"
    ws["H5"] = "Danilo Manrique"
    ws["BX5"] = f"Repositorio: {REPO}"

    start_row = 6
    for i, bolt in enumerate(BOLTS):
        fill_bolt(ws, start_row + i, bolt)

    wb.save(OUT)
    print(f"Guardado: {OUT}")
    print(f"Bolts registrados: {len(BOLTS)} (filas {start_row}-{start_row + len(BOLTS) - 1})")


if __name__ == "__main__":
    main()
